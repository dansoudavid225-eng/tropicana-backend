from rest_framework import generics, status, permissions
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate
from django.core.mail import send_mail as _send_mail_original

def send_mail(*args, **kwargs):
    kwargs['fail_silently'] = True
    try:
        _send_mail_original(*args, **kwargs)
    except Exception:
        pass
from django.conf import settings
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.db import models
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
import datetime
import logging
import hashlib
import hmac
import json

logger_securite = logging.getLogger('api.securite')

try:
    from google.oauth2 import id_token
    from google.auth.transport import requests as google_requests
    GOOGLE_AUTH_AVAILABLE = True
except ImportError:
    GOOGLE_AUTH_AVAILABLE = False

from .models import Utilisateur, Produit, Commande, Temoignage, MessageContact, ResetPasswordToken, PanierSauvegarde, LogConnexion
from .serializers import (
    InscriptionSerializer, UtilisateurSerializer,
    ModifierProfilSerializer, ModifierMotDePasseSerializer,
    GoogleAuthSerializer,
    ProduitSerializer,
    CommandeCreerSerializer, CommandeSerializer,
    TemoignageCreerSerializer, TemoignageSerializer,
    MessageContactSerializer,
)


# ─── Helpers ──────────────────────────────────────────────────────────────────

def tokens_pour_utilisateur(user):
    from .serializers import UtilisateurSerializer
    refresh = RefreshToken.for_user(user)
    return {
        'refresh':     str(refresh),
        'access':      str(refresh.access_token),
        'utilisateur': UtilisateurSerializer(user).data,
    }


def envoyer_confirmation_commande(commande):
    """Envoie un email de confirmation au client après sa commande."""
    if not commande.email_client:
        return
    try:
        lignes_txt = '\n'.join([
            f"   • {l.produit.nom if l.produit else 'Produit'} x{l.quantite} — {getattr(l, 'prix_unitaire', '')} FCFA"
            for l in commande.lignes.all()
        ])
        send_mail(
            subject=f'[Tropicana Pio Pio] ✅ Commande #{commande.pk} reçue !',
            message=(
                f"Bonjour {commande.nom_client},\n\n"
                f"Nous avons bien reçu votre commande #{commande.pk}.\n"
                f"Notre équipe vous contactera sous 2h pour confirmer la livraison.\n\n"
                f"📋 Détail de votre commande :\n{lignes_txt}\n\n"
                f"💰 Total : {commande.total} FCFA\n"
                f"🏠 Livraison : {commande.ville_livraison}\n"
                f"💳 Paiement : {commande.get_mode_paiement_display()}\n\n"
                f"Pour toute question :\n"
                f"📞 +229 01 95 96 77 62\n"
                f"💬 WhatsApp : wa.me/2290195967762\n\n"
                f"Merci de votre confiance 🌿\n"
                f"\n🔍 Suivre votre commande : {settings.FRONTEND_URL}/suivi-commande\n"
                f"(Numéro de commande : #{commande.pk} — Email : {commande.email_client})\n\n"
                f"— L'équipe Tropicana Pio Pio"
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[commande.email_client],
            fail_silently=True,
        )
        # Notification interne
        send_mail(
            subject=f'🛒 Nouvelle commande #{commande.pk} — {commande.total} FCFA',
            message=(
                f"Client    : {commande.nom_client}\n"
                f"Email     : {commande.email_client}\n"
                f"Téléphone : {commande.telephone_client}\n"
                f"Ville     : {commande.ville_livraison}\n"
                f"Total     : {commande.total} FCFA\n"
                f"Paiement  : {commande.get_mode_paiement_display()}\n\n"
                f"Produits :\n{lignes_txt}"
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[settings.DEFAULT_FROM_EMAIL],
            fail_silently=True,
        )
    except Exception as e:
        logger_securite.error(f'Email confirmation commande #{commande.pk} : {e}')


def envoyer_notification_statut(commande):
    """Envoie un email au client quand le statut de sa commande change."""
    if not commande.email_client:
        return

    messages_statut = {
        'confirmee':    ('✅ Commande confirmée', 'Votre commande a été confirmée par notre équipe. Nous préparons votre colis.'),
        'en_livraison': ('🚚 En cours de livraison', 'Votre commande est en route ! Notre livreur vous contactera bientôt.'),
        'livree':       ('📦 Commande livrée', 'Votre commande a été livrée. Merci de votre confiance !'),
        'annulee':      ('❌ Commande annulée', 'Votre commande a été annulée. Contactez-nous pour plus d\'informations.'),
    }

    if commande.statut not in messages_statut:
        return

    sujet_emoji, message_statut = messages_statut[commande.statut]
    send_mail(
        subject=f"[Tropicana Pio Pio] {sujet_emoji} — Commande #{commande.pk}",
        message=(
            f"Bonjour {commande.nom_client},\n\n"
            f"{message_statut}\n\n"
            f"📋 Commande #{commande.pk} — Total : {commande.total} FCFA\n"
            f"🏠 Livraison : {commande.ville_livraison}\n\n"
            f"📞 +229 01 95 96 77 62\n"
            f"💬 WhatsApp : wa.me/2290195967762\n\n"
            f"— L'équipe Tropicana Pio Pio"
        ),
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[commande.email_client],
        fail_silently=True,
    )


# ─── Auth ─────────────────────────────────────────────────────────────────────

def _get_client_ip(request):
    x_forwarded = request.META.get('HTTP_X_FORWARDED_FOR', '')
    if x_forwarded:
        return x_forwarded.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', '')


class InscriptionView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        serializer = InscriptionSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            try:
                send_mail(
                    subject='🌿 Bienvenue chez Tropicana Pio Pio !',
                    message=(
                        f"Bonjour {user.prenom},\n\n"
                        f"Votre compte Tropicana Pio Pio a été créé avec succès.\n"
                        f"Découvrez notre boutique : https://tropicanapiopio.com/boutique\n\n"
                        f"— L'équipe Tropicana Pio Pio 🌿"
                    ),
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[user.email],
                    fail_silently=True,
                )
            except Exception:
                pass
            return Response(
                {'message': 'Compte créé avec succès.', **tokens_pour_utilisateur(user)},
                status=status.HTTP_201_CREATED
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ConnexionView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        email       = request.data.get('email', '').lower().strip()
        mot_de_passe = request.data.get('mot_de_passe', '')
        ip          = _get_client_ip(request)

        if not email or not mot_de_passe:
            return Response({'detail': 'Email et mot de passe requis.'}, status=status.HTTP_400_BAD_REQUEST)

        # Vérification blacklist email/IP
        if Blacklist.objects.filter(valeur__in=[email, ip]).exists():
            return Response({'detail': 'Accès restreint.'}, status=status.HTTP_403_FORBIDDEN)

        user = authenticate(request, username=email, password=mot_de_passe)

        if user:
            LogConnexion.objects.create(email=email, ip=ip, resultat='succes',
                                        user_agent=request.META.get('HTTP_USER_AGENT', '')[:255])
            return Response({'message': 'Connexion réussie.', **tokens_pour_utilisateur(user)})
        else:
            LogConnexion.objects.create(email=email, ip=ip, resultat='echec',
                                        user_agent=request.META.get('HTTP_USER_AGENT', '')[:255])
            logger_securite.warning(f"Tentative connexion échouée — email: {email} — IP: {ip}")
            return Response({'detail': 'Identifiants incorrects.'}, status=status.HTTP_401_UNAUTHORIZED)


class DeconnexionView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        try:
            token = RefreshToken(request.data.get('refresh'))
            token.blacklist()
        except Exception:
            pass
        return Response({'message': 'Déconnexion réussie.'})


class GoogleAuthView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        if not GOOGLE_AUTH_AVAILABLE:
            return Response({'detail': 'Google Auth non configuré.'}, status=status.HTTP_501_NOT_IMPLEMENTED)
        serializer = GoogleAuthSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        id_token_str = serializer.validated_data['id_token']
        try:
            idinfo = id_token.verify_oauth2_token(
                id_token_str, google_requests.Request(), settings.GOOGLE_CLIENT_ID
            )
            email     = idinfo.get('email', '').lower()
            prenom    = idinfo.get('given_name', '')
            nom       = idinfo.get('family_name', '')
            if not email:
                return Response({'detail': 'Email manquant dans le token Google.'}, status=400)
            user, _ = Utilisateur.objects.get_or_create(
                email=email,
                defaults={'prenom': prenom, 'nom': nom, 'is_active': True}
            )
            return Response({'message': 'Connexion Google réussie.', **tokens_pour_utilisateur(user)})
        except ValueError as e:
            return Response({'detail': f'Token Google invalide : {e}'}, status=status.HTTP_401_UNAUTHORIZED)


class DemandeResetMotDePasseView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        email = request.data.get('email', '').strip().lower()
        if not email:
            return Response({'detail': 'Email requis.'}, status=400)
        # Réponse identique qu'il existe ou pas (évite l'énumération d'emails)
        try:
            user = Utilisateur.objects.get(email=email)
            token_obj, _ = ResetPasswordToken.objects.get_or_create(
                utilisateur=user,
                defaults={
                    'token': get_random_string(64),
                    'expire_le': timezone.now() + datetime.timedelta(hours=1),
                    'utilise': False,
                }
            )
            token_obj.token    = get_random_string(64)
            token_obj.expire_le = timezone.now() + datetime.timedelta(hours=1)
            token_obj.utilise  = False
            token_obj.save()
            lien = f"{settings.FRONTEND_URL}/reinitialiser-mot-de-passe?token={token_obj.token}"
            send_mail(
                subject='[Tropicana Pio Pio] Réinitialisation de votre mot de passe',
                message=(
                    f"Bonjour {user.prenom},\n\n"
                    f"Cliquez sur ce lien pour réinitialiser votre mot de passe (valable 1h) :\n{lien}\n\n"
                    f"Si vous n'avez pas fait cette demande, ignorez cet email.\n\n"
                    f"— L'équipe Tropicana Pio Pio"
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=True,
            )
        except Utilisateur.DoesNotExist:
            pass
        return Response({'detail': 'Si cet email existe, un lien de réinitialisation a été envoyé.'})


class VerifierTokenResetView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        token = request.query_params.get('token', '')
        try:
            t = ResetPasswordToken.objects.get(token=token, utilise=False)
            if t.expire_le < timezone.now():
                return Response({'detail': 'Lien expiré.'}, status=400)
            return Response({'detail': 'Token valide.'})
        except ResetPasswordToken.DoesNotExist:
            return Response({'detail': 'Token invalide.'}, status=400)


class ConfirmerResetMotDePasseView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        token        = request.data.get('token', '')
        nouveau_mdp  = request.data.get('nouveau_mot_de_passe', '')
        if not token or not nouveau_mdp or len(nouveau_mdp) < 8:
            return Response({'detail': 'Token et mot de passe (8 caractères min) requis.'}, status=400)
        try:
            t = ResetPasswordToken.objects.get(token=token, utilise=False)
            if t.expire_le < timezone.now():
                return Response({'detail': 'Lien expiré.'}, status=400)
            t.utilisateur.set_password(nouveau_mdp)
            t.utilisateur.save()
            t.utilise = True
            t.save()
            return Response({'detail': 'Mot de passe mis à jour avec succès.'})
        except ResetPasswordToken.DoesNotExist:
            return Response({'detail': 'Token invalide.'}, status=400)


class ProfilView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        return Response(UtilisateurSerializer(request.user).data)

    def patch(self, request):
        serializer = ModifierProfilSerializer(request.user, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)


class ModifierMotDePasseView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        serializer = ModifierMotDePasseSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            request.user.set_password(serializer.validated_data['nouveau_mot_de_passe'])
            request.user.save()
            return Response({'detail': 'Mot de passe modifié avec succès.'})
        return Response(serializer.errors, status=400)


class PanierView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        panier, _ = PanierSauvegarde.objects.get_or_create(utilisateur=request.user)
        return Response(panier.donnees or {'lignes': []})

    def post(self, request):
        panier, _ = PanierSauvegarde.objects.get_or_create(utilisateur=request.user)
        panier.donnees = request.data
        panier.save()
        return Response({'detail': 'Panier sauvegardé.'})

    def delete(self, request):
        PanierSauvegarde.objects.filter(utilisateur=request.user).update(donnees={'lignes': []})
        return Response({'detail': 'Panier vidé.'})


# ─── Produits ─────────────────────────────────────────────────────────────────

class ProduitsListView(generics.ListAPIView):
    permission_classes = [permissions.AllowAny]
    serializer_class   = ProduitSerializer

    def get_queryset(self):
        return Produit.objects.filter(disponible=True)

    def get_serializer_context(self):
        return {'request': self.request}


class ProduitDetailView(generics.RetrieveAPIView):
    permission_classes = [permissions.AllowAny]
    serializer_class   = ProduitSerializer
    queryset           = Produit.objects.filter(disponible=True)
    lookup_field       = 'slug'

    def get_serializer_context(self):
        return {'request': self.request}


# ─── Commandes ────────────────────────────────────────────────────────────────

class CommandeCreerView(APIView):
    """Crée une commande. Si mode_paiement=fedapay, initie la transaction FedaPay."""
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        serializer = CommandeCreerSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # ✅ Vérification du stock AVANT de créer la commande
        lignes_data = serializer.validated_data.get('lignes', [])
        for ligne in lignes_data:
            produit  = ligne.get('produit')
            quantite = ligne.get('quantite', 1)
            if produit and produit.stock > 0 and produit.stock < quantite:
                return Response(
                    {'detail': f'Stock insuffisant pour "{produit.nom}". Stock disponible : {produit.stock}.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # ✅ Vérification code promo (si fourni)
        code_promo_str = request.data.get('code_promo', '').strip().upper()
        reduction = 0
        code_promo_obj = None
        if code_promo_str:
            from .models import CodePromo
            try:
                cp = CodePromo.objects.get(code=code_promo_str, actif=True)
                if cp.est_valide():
                    reduction = cp.calculer_reduction(serializer.validated_data.get('total', 0))
                    code_promo_obj = cp
                else:
                    return Response({'detail': 'Code promo expiré ou limite atteinte.'}, status=400)
            except CodePromo.DoesNotExist:
                return Response({'detail': 'Code promo invalide.'}, status=400)

        commande = serializer.save()

        # Appliquer réduction si code promo valide
        if code_promo_obj and reduction > 0:
            commande.total = max(0, commande.total - reduction)
            commande.code_promo = code_promo_str
            commande.save(update_fields=['total', 'code_promo'])
            code_promo_obj.nb_utilisations += 1
            code_promo_obj.save(update_fields=['nb_utilisations'])

        if request.user.is_authenticated:
            commande.utilisateur = request.user
            if not commande.nom_client:
                commande.nom_client = request.user.nom_complet
            if not commande.email_client:
                commande.email_client = request.user.email
            commande.save(update_fields=['utilisateur', 'nom_client', 'email_client'])

        # ✅ Décrémenter le stock
        for ligne in commande.lignes.select_related('produit'):
            p = ligne.produit
            if p and p.stock > 0:
                p.stock = max(0, p.stock - ligne.quantite)
                if p.stock == 0:
                    p.disponible = False
                p.save(update_fields=['stock', 'disponible'])

        # Vérifier alertes stock
        try:
            verifier_alertes_stock()
        except Exception as e:
            logger_securite.error(f'Alerte stock error: {e}')

        # Notification email à l'admin (sauf FedaPay : on attend la confirmation du paiement)
        if commande.mode_paiement != 'fedapay':
            try:
                lignes_txt = '\n'.join(
                    f"- {l.produit.nom if l.produit else '?'} x{l.quantite}"
                    for l in commande.lignes.select_related('produit').all()
                )
                send_mail(
                    subject=f'Nouvelle commande #{commande.pk} - {commande.total} FCFA',
                    message=(
                        f"Nouvelle commande recue sur Tropicana Pio Pio !\n\n"
                        f"Commande : #{commande.pk}\n"
                        f"Client : {commande.nom_client or '-'}\n"
                        f"Telephone : {commande.telephone_client or '-'}\n"
                        f"Ville : {commande.ville_livraison or '-'}\n"
                        f"Mode de paiement : {commande.mode_paiement}\n"
                        f"Total : {commande.total} FCFA\n\n"
                        f"Produits :\n{lignes_txt}\n\n"
                        f"Voir dans l'admin : https://tropicana-api.onrender.com/admin/api/commande/{commande.pk}/change/"
                    ),
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=['tropicanapiopio.officiel@gmail.com'],
                    fail_silently=True,
                )
            except Exception as e:
                logger_securite.error(f'Email notification commande error: {e}')

        # Vider le panier backend
        if request.user.is_authenticated:
            PanierSauvegarde.objects.filter(utilisateur=request.user).update(donnees={'lignes': []})

        # ── Paiement FedaPay ──────────────────────────────────────────────
        if commande.mode_paiement == 'fedapay' and settings.FEDAPAY_SECRET_KEY:
            try:
                import requests as req_lib
                frontend_url = getattr(settings, 'FRONTEND_URL', 'https://tropicana-pio-pio.netlify.app')
                fedapay_env  = getattr(settings, 'FEDAPAY_ENV', 'live')
                base_url     = 'https://api.fedapay.com' if fedapay_env == 'live' else 'https://sandbox-api.fedapay.com'
                headers      = {
                    'Authorization': f'Bearer {settings.FEDAPAY_SECRET_KEY}',
                    'Content-Type':  'application/json',
                }
                payload = {
                    'description': f'Commande Tropicana Pio Pio #{commande.pk}',
                    'amount':      int(commande.total),
                    'currency':    {'iso': 'XOF'},
                    'callback_url': f'{frontend_url}/paiement/retour?commande={commande.pk}&email={commande.email_client}',
                    'customer': {
                        'firstname': commande.nom_client.split()[0] if commande.nom_client else 'Client',
                        'lastname':  ' '.join(commande.nom_client.split()[1:]) if commande.nom_client else '',
                        'email':     commande.email_client or '',
                        'phone_number': {
                            'number':  commande.telephone_client.replace(' ', '').replace('+229', '').replace('+', '') if commande.telephone_client else '',
                            'country': 'BJ',
                        }
                    }
                }
                r = req_lib.post(f'{base_url}/v1/transactions', json=payload, headers=headers, timeout=30)
                r.raise_for_status()
                r_json = r.json()
                transaction_obj = (
                    r_json.get('v1/transaction') or
                    r_json.get('transaction') or
                    r_json
                )
                transaction_id = transaction_obj['id']
                commande.fedapay_ref = str(transaction_id)
                commande.save(update_fields=['fedapay_ref'])
                r2 = req_lib.post(f'{base_url}/v1/transactions/{transaction_id}/token', headers=headers, timeout=30)
                r2.raise_for_status()
                r2_json = r2.json()
                fedapay_url = r2_json.get('url') or r2_json.get('token', {}).get('url', '')
                return Response({
                    'message':     'Commande créée. Redirigez vers FedaPay pour payer.',
                    'commande_id': commande.pk,
                    'total':       commande.total,
                    'reduction':   reduction,
                    'statut':      commande.statut,
                    'fedapay_url': fedapay_url,
                }, status=status.HTTP_201_CREATED)
            except Exception as e:
                logger_securite.error(f'FedaPay error commande #{commande.pk}: {e}')
                return Response({
                    'message':     'Commande enregistrée. Le paiement en ligne est temporairement indisponible — notre équipe vous contactera.',
                    'commande_id': commande.pk,
                    'total':       commande.total,
                    'reduction':   reduction,
                    'statut':      commande.statut,
                    'fedapay_url': None,
                }, status=status.HTTP_201_CREATED)

        # ── Autres modes de paiement ──────────────────────────────────────
        envoyer_confirmation_commande(commande)
        return Response({
            'message':     'Commande enregistrée ! Notre équipe vous contactera sous 2h.',
            'commande_id': commande.pk,
            'total':       commande.total,
            'reduction':   reduction,
            'statut':      commande.statut,
        }, status=status.HTTP_201_CREATED)


@method_decorator(csrf_exempt, name='dispatch')
class FedapayWebhookView(APIView):
    """Reçoit les callbacks FedaPay et met à jour le statut de la commande."""
    permission_classes = [permissions.AllowAny]
    throttle_scope     = 'webhook'

    def post(self, request):
        # Vérification signature webhook (uniquement si FEDAPAY_WEBHOOK_SECRET est défini)
        secret = getattr(settings, 'FEDAPAY_WEBHOOK_SECRET', '')
        if secret:
            sig  = request.META.get('HTTP_X_FEDAPAY_SIGNATURE', '')
            body = request.body
            mac      = hmac.new(secret.encode(), body, hashlib.sha256)
            expected = mac.hexdigest()
            if not hmac.compare_digest(sig, expected):
                return Response({'detail': 'Signature invalide'}, status=status.HTTP_403_FORBIDDEN)
        # Pas de secret configuré → on accepte tous les webhooks FedaPay (mode sans signature)

        try:
            payload = json.loads(request.body)
            logger_securite.error(f'FedaPay webhook payload: {payload}')
            event = payload.get('name', '')

            # FedaPay peut envoyer la transaction sous differentes cles selon la version d'API
            transaction_data = (
                payload.get('data', {}).get('object', {}) or
                payload.get('entity', {}) or
                payload.get('data', {}) or
                {}
            )
            transaction_id   = str(transaction_data.get('id', '') or '')
            transaction_status = transaction_data.get('status', '')

            if not transaction_id:
                logger_securite.error(f'FedaPay webhook: transaction_id manquant dans {payload}')
                return Response({'detail': 'Donnees manquantes'}, status=status.HTTP_400_BAD_REQUEST)

            try:
                commande = Commande.objects.get(fedapay_ref=transaction_id)
            except Commande.DoesNotExist:
                return Response({'detail': 'Commande introuvable'}, status=status.HTTP_404_NOT_FOUND)

            if transaction_status in ('approved', 'transferred') or event == 'transaction.approved':
                commande.payee  = True
                commande.statut = 'confirmee'
                commande.save(update_fields=['payee', 'statut'])
                envoyer_confirmation_commande(commande)
                try:
                    lignes_txt = '\n'.join(
                        f"- {l.produit.nom if l.produit else '?'} x{l.quantite}"
                        for l in commande.lignes.select_related('produit').all()
                    )
                    send_mail(
                        subject=f'Paiement confirme - commande #{commande.pk} - {commande.total} FCFA',
                        message=(
                            f"Paiement confirme sur Tropicana Pio Pio !\n\n"
                            f"Commande : #{commande.pk}\n"
                            f"Client : {commande.nom_client or '-'}\n"
                            f"Telephone : {commande.telephone_client or '-'}\n"
                            f"Ville : {commande.ville_livraison or '-'}\n"
                            f"Mode de paiement : {commande.mode_paiement}\n"
                            f"Total : {commande.total} FCFA\n\n"
                            f"Produits :\n{lignes_txt}\n\n"
                            f"Voir dans l'admin : https://tropicana-api.onrender.com/admin/api/commande/{commande.pk}/change/"
                        ),
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=['tropicanapiopio.officiel@gmail.com'],
                        fail_silently=True,
                    )
                except Exception as e:
                    logger_securite.error(f'Email notification paiement confirme error: {e}')

            elif transaction_status in ('declined', 'canceled') or event in ('transaction.declined', 'transaction.canceled'):
                commande.statut = 'annulee'
                commande.save(update_fields=['statut'])

            return Response({'detail': 'OK'}, status=status.HTTP_200_OK)

        except Exception as e:
            logger_securite.error(f'FedaPay webhook error: {e}')
            return Response({'detail': 'Erreur serveur'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class MesCommandesView(generics.ListAPIView):
    serializer_class   = CommandeSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Commande.objects.filter(utilisateur=self.request.user).prefetch_related('lignes__produit').order_by('-date_commande')


class CommandeDetailView(generics.RetrieveAPIView):
    serializer_class   = CommandeSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_object(self):
        try:
            return Commande.objects.prefetch_related('lignes__produit').get(
                pk=self.kwargs['pk'], utilisateur=self.request.user
            )
        except Commande.DoesNotExist:
            from rest_framework.exceptions import NotFound
            raise NotFound('Commande introuvable.')


# ─── Témoignages ──────────────────────────────────────────────────────────────

class TemoignagesListView(generics.ListCreateAPIView):
    permission_classes = [permissions.AllowAny]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return TemoignageCreerSerializer
        return TemoignageSerializer

    def get_queryset(self):
        return Temoignage.objects.filter(approuve=True).order_by('-date_creation')

    def get_serializer_context(self):
        return {'request': self.request}


# ─── Contact ──────────────────────────────────────────────────────────────────

class ContactView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        # Vérification blacklist
        email = request.data.get('email', '').strip().lower()
        ip    = _get_client_ip(request)
        if Blacklist.objects.filter(valeur__in=[email, ip]).exists():
            return Response({'detail': 'Votre accès a été restreint.'}, status=status.HTTP_403_FORBIDDEN)

        serializer = MessageContactSerializer(data=request.data)
        if serializer.is_valid():
            msg = serializer.save()
            send_mail(
                subject=f'[Tropicana Pio Pio] Nouveau message — {msg.sujet}',
                message=(
                    f"De : {msg.nom} <{msg.email}>\n"
                    f"Sujet : {msg.sujet}\n\n"
                    f"{msg.message}"
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[settings.DEFAULT_FROM_EMAIL],
                fail_silently=True,
            )
            return Response({'detail': 'Message envoyé. Nous vous répondrons sous 24h.'}, status=201)
        return Response(serializer.errors, status=400)


# ─── Vues Admin ───────────────────────────────────────────────────────────────

class AdminCommandesView(generics.ListAPIView):
    """Toutes les commandes — admin seulement — avec filtres, pagination et export CSV/Excel."""
    serializer_class   = CommandeSerializer
    permission_classes = [permissions.IsAdminUser]

    def get_queryset(self):
        qs     = Commande.objects.all().prefetch_related('lignes__produit').order_by('-date_commande')
        statut = self.request.query_params.get('statut')
        ville  = self.request.query_params.get('ville')
        search = self.request.query_params.get('search')
        date_debut = self.request.query_params.get('date_debut')
        date_fin   = self.request.query_params.get('date_fin')
        if statut:
            qs = qs.filter(statut=statut)
        if ville:
            qs = qs.filter(ville_livraison__icontains=ville)
        if search:
            qs = qs.filter(
                models.Q(nom_client__icontains=search) |
                models.Q(email_client__icontains=search) |
                models.Q(telephone_client__icontains=search)
            )
        if date_debut:
            qs = qs.filter(date_commande__date__gte=date_debut)
        if date_fin:
            qs = qs.filter(date_commande__date__lte=date_fin)
        return qs

    def list(self, request, *args, **kwargs):
        fmt = request.query_params.get('format', '')
        qs  = self.get_queryset()
        if fmt == 'csv':
            return self._export_csv(qs)
        if fmt == 'excel':
            return self._export_excel(qs)

        # Pagination
        page     = int(request.query_params.get('page', 1))
        per_page = int(request.query_params.get('per_page', 20))
        total    = qs.count()
        start    = (page - 1) * per_page
        data     = self.get_serializer(qs[start:start + per_page], many=True).data
        return Response({
            'count':    total,
            'page':     page,
            'pages':    (total + per_page - 1) // per_page,
            'per_page': per_page,
            'results':  data,
        })

    def _export_csv(self, qs):
        import csv
        from django.http import HttpResponse
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="commandes_tropicana.csv"'
        writer = csv.writer(response)
        writer.writerow(['ID', 'Date', 'Client', 'Email', 'Téléphone', 'Ville',
                         'Statut', 'Mode paiement', 'FedaPay ref', 'Payée', 'Total (FCFA)', 'Produits'])
        for cmd in qs:
            produits = ' | '.join([f"{l.produit.nom if l.produit else 'Produit'} x{l.quantite}" for l in cmd.lignes.all()])
            writer.writerow([
                cmd.id,
                cmd.date_commande.strftime('%d/%m/%Y %H:%M'),
                cmd.nom_client,
                cmd.email_client,
                cmd.telephone_client,
                cmd.ville_livraison,
                cmd.get_statut_display(),
                cmd.get_mode_paiement_display(),
                cmd.fedapay_ref or '',
                'Oui' if cmd.payee else 'Non',
                cmd.total,
                produits,
            ])
        return response

    def _export_excel(self, qs):
        """Export Excel (.xlsx) avec openpyxl."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import HttpResponse

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Commandes'

            # En-têtes
            entetes = ['ID', 'Date', 'Client', 'Email', 'Téléphone', 'Ville',
                       'Statut', 'Mode paiement', 'FedaPay ref', 'Payée', 'Total (FCFA)', 'Produits']
            vert_fonce = 'FF2D6A4F'
            for col, titre in enumerate(entetes, 1):
                cell = ws.cell(row=1, column=col, value=titre)
                cell.font      = Font(bold=True, color='FFFFFFFF')
                cell.fill      = PatternFill('solid', fgColor=vert_fonce)
                cell.alignment = Alignment(horizontal='center')

            # Données
            for row_idx, cmd in enumerate(qs, 2):
                produits = ' | '.join([f"{l.produit.nom if l.produit else 'Produit'} x{l.quantite}" for l in cmd.lignes.all()])
                valeurs = [
                    cmd.id,
                    cmd.date_commande.strftime('%d/%m/%Y %H:%M'),
                    cmd.nom_client,
                    cmd.email_client,
                    cmd.telephone_client,
                    cmd.ville_livraison,
                    cmd.get_statut_display(),
                    cmd.get_mode_paiement_display(),
                    cmd.fedapay_ref or '',
                    'Oui' if cmd.payee else 'Non',
                    int(cmd.total),
                    produits,
                ]
                for col, val in enumerate(valeurs, 1):
                    ws.cell(row=row_idx, column=col, value=val)
                # Alterner couleurs lignes
                if row_idx % 2 == 0:
                    for col in range(1, len(entetes) + 1):
                        ws.cell(row=row_idx, column=col).fill = PatternFill('solid', fgColor='FFF0F4F1')

            # Largeur colonnes automatique
            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="commandes_tropicana.xlsx"'
            wb.save(response)
            return response

        except ImportError:
            # openpyxl non installé → fallback CSV
            return self._export_csv(qs)


class AdminCommandeDetailView(generics.RetrieveUpdateAPIView):
    serializer_class   = CommandeSerializer
    permission_classes = [permissions.IsAdminUser]
    queryset           = Commande.objects.all().prefetch_related('lignes__produit')

    def partial_update(self, request, *args, **kwargs):
        instance      = self.get_object()
        statut        = request.data.get('statut')
        ancien_statut = instance.statut
        if statut and statut in dict(Commande.STATUT_CHOICES):
            instance.statut = statut
            instance.save(update_fields=['statut'])
            if statut != ancien_statut:
                # Enregistrer l'historique
                HistoriqueCommande.objects.create(
                    commande=instance,
                    ancien_statut=ancien_statut,
                    nouveau_statut=statut,
                    note=request.data.get('note', ''),
                    modifie_par=request.user if request.user.is_authenticated else None,
                )
                envoyer_notification_statut(instance)
        return Response(CommandeSerializer(instance).data)


# ─── Stats analytiques admin ──────────────────────────────────────────────────

class AdminStatsView(APIView):
    """Statistiques avancées pour le dashboard admin React."""
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        from django.db.models import Sum, Count
        from .models import NewsletterAbonne, ArticleBlog

        now        = timezone.now()
        debut_mois = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        # Stats générales
        total_commandes      = Commande.objects.count()
        commandes_attente    = Commande.objects.filter(statut='en_attente').count()
        total_clients        = Utilisateur.objects.filter(is_staff=False).count()
        newsletter_abonnes   = NewsletterAbonne.objects.filter(actif=True).count()
        produits_dispo       = Produit.objects.filter(disponible=True).count()
        articles_publie      = ArticleBlog.objects.filter(publie=True).count()
        messages_non_lus     = MessageContact.objects.filter(lu=False).count()

        ca_total = Commande.objects.filter(statut='livree').aggregate(s=Sum('total'))['s'] or 0
        ca_mois  = Commande.objects.filter(statut='livree', date_commande__gte=debut_mois).aggregate(s=Sum('total'))['s'] or 0

        # Commandes par statut
        par_statut = {
            item['statut']: item['nb']
            for item in Commande.objects.values('statut').annotate(nb=Count('id'))
        }

        # CA des 12 derniers mois
        mois_fr = ['Jan','Fév','Mar','Avr','Mai','Jun','Jul','Aoû','Sep','Oct','Nov','Déc']
        ca_par_mois = []
        for i in range(11, -1, -1):
            d      = now - datetime.timedelta(days=30 * i)
            debut  = d.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            fin    = (debut.replace(month=debut.month % 12 + 1, day=1)
                      if debut.month < 12
                      else debut.replace(year=debut.year + 1, month=1, day=1))
            ca = Commande.objects.filter(
                statut='livree', date_commande__gte=debut, date_commande__lt=fin
            ).aggregate(s=Sum('total'))['s'] or 0
            nb = Commande.objects.filter(
                date_commande__gte=debut, date_commande__lt=fin
            ).count()
            ca_par_mois.append({
                'mois':     mois_fr[debut.month - 1],
                'ca':       int(ca),
                'commandes': nb,
            })

        # Top produits
        from django.db.models import Sum as DSum
        from .models import LigneCommande
        top_produits = list(
            LigneCommande.objects.values('produit_nom')
            .annotate(total_vendu=DSum('quantite'), ca=DSum(models.F('quantite') * models.F('prix_unitaire')))
            .order_by('-total_vendu')[:5]
        )

        return Response({
            'total_commandes':    total_commandes,
            'commandes_attente':  commandes_attente,
            'total_clients':      total_clients,
            'newsletter_abonnes': newsletter_abonnes,
            'produits_dispo':     produits_dispo,
            'articles_publie':    articles_publie,
            'messages_non_lus':   messages_non_lus,
            'ca_total':           int(ca_total),
            'ca_mois':            int(ca_mois),
            'par_statut':         par_statut,
            'ca_par_mois':        ca_par_mois,
            'top_produits':       top_produits,
        })


# ─── Code Promo ───────────────────────────────────────────────────────────────

class AdminCodePromoView(generics.ListCreateAPIView):
    """Gestion des codes promo — admin seulement."""
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        from .models import CodePromo
        promos = CodePromo.objects.all().order_by('-date_creation')
        data = [{
            'id':             p.id,
            'code':           p.code,
            'type_reduction': p.type_reduction,
            'valeur':         p.valeur,
            'nb_utilisations': p.nb_utilisations,
            'limite_utilisations': p.limite_utilisations,
            'date_expiration': p.date_expiration.isoformat() if p.date_expiration else None,
            'actif':          p.actif,
            'valide':         p.est_valide(),
            'date_creation':  p.date_creation.isoformat(),
        } for p in promos]
        return Response(data)

    def post(self, request):
        from .models import CodePromo
        code   = request.data.get('code', '').strip().upper()
        valeur = request.data.get('valeur', 0)
        type_r = request.data.get('type_reduction', 'pourcentage')
        limite = request.data.get('limite_utilisations', None)
        expiry = request.data.get('date_expiration', None)

        if not code:
            return Response({'detail': 'Code requis.'}, status=400)
        if CodePromo.objects.filter(code=code).exists():
            return Response({'detail': 'Ce code existe déjà.'}, status=400)

        promo = CodePromo.objects.create(
            code=code,
            type_reduction=type_r,
            valeur=valeur,
            limite_utilisations=limite,
            date_expiration=expiry,
        )
        return Response({'detail': 'Code promo créé.', 'id': promo.id}, status=201)


class AdminCodePromoDetailView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def patch(self, request, pk):
        from .models import CodePromo
        try:
            promo = CodePromo.objects.get(pk=pk)
        except CodePromo.DoesNotExist:
            return Response({'detail': 'Introuvable.'}, status=404)
        for champ in ['actif', 'valeur', 'limite_utilisations', 'date_expiration']:
            if champ in request.data:
                setattr(promo, champ, request.data[champ])
        promo.save()
        return Response({'detail': 'Mis à jour.'})

    def delete(self, request, pk):
        from .models import CodePromo
        CodePromo.objects.filter(pk=pk).delete()
        return Response({'detail': 'Supprimé.'})


class ValiderCodePromoView(APIView):
    """Vérifie un code promo (public) et retourne la réduction applicable."""
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        from .models import CodePromo
        code  = request.data.get('code', '').strip().upper()
        total = request.data.get('total', 0)
        if not code:
            return Response({'detail': 'Code requis.'}, status=400)
        try:
            cp = CodePromo.objects.get(code=code, actif=True)
            if not cp.est_valide():
                return Response({'detail': 'Code expiré ou limite atteinte.'}, status=400)
            reduction = cp.calculer_reduction(total)
            return Response({
                'valide':         True,
                'type_reduction': cp.type_reduction,
                'valeur':         cp.valeur,
                'reduction':      reduction,
                'nouveau_total':  max(0, int(total) - reduction),
            })
        except CodePromo.DoesNotExist:
            return Response({'detail': 'Code invalide.'}, status=400)


# ─── Messages, Config, Produits, Témoignages, Utilisateurs ───────────────────

class AdminMessagesView(generics.ListAPIView):
    serializer_class   = MessageContactSerializer
    permission_classes = [permissions.IsAdminUser]
    queryset           = MessageContact.objects.all().order_by('-date_envoi')


class AdminMessageDetailView(generics.RetrieveUpdateAPIView):
    serializer_class   = MessageContactSerializer
    permission_classes = [permissions.IsAdminUser]
    queryset           = MessageContact.objects.all()


class AdminSiteConfigView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def _get_path(self):
        import os
        return os.path.join(settings.BASE_DIR, 'site_config.json')

    def get(self, request):
        import os
        path = self._get_path()
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as f:
                config = json.load(f)
        else:
            config = {}
        return Response(config)

    def post(self, request):
        path = self._get_path()
        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(request.data, f, ensure_ascii=False, indent=2)
            return Response({'message': 'Configuration sauvegardée.', 'config': request.data})
        except Exception as e:
            return Response({'detail': str(e)}, status=400)


class AdminProduitsView(generics.ListCreateAPIView):
    serializer_class   = ProduitSerializer
    permission_classes = [permissions.IsAdminUser]
    queryset           = Produit.objects.all()

    def get_serializer_context(self):
        return {'request': self.request}


class AdminProduitDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class   = ProduitSerializer
    permission_classes = [permissions.IsAdminUser]
    queryset           = Produit.objects.all()

    def get_serializer_context(self):
        return {'request': self.request}


class AdminTemoignagesView(generics.ListAPIView):
    serializer_class   = TemoignageSerializer
    permission_classes = [permissions.IsAdminUser]

    def get_queryset(self):
        return Temoignage.objects.all().order_by('-date_creation')


class AdminTemoignageDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class   = TemoignageSerializer
    permission_classes = [permissions.IsAdminUser]
    queryset           = Temoignage.objects.all()

    def get_serializer_context(self):
        return {'request': self.request}

    def patch(self, request, *args, **kwargs):
        t = self.get_object()
        t.approuve = request.data.get('approuve', t.approuve)
        t.save(update_fields=['approuve'])
        return Response(TemoignageSerializer(t, context={'request': request}).data)


class AdminUtilisateursView(generics.ListAPIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request, *args, **kwargs):
        users = Utilisateur.objects.all().order_by('-date_inscription')
        data = [{
            'id':               u.id,
            'prenom':           u.prenom,
            'nom':              u.nom,
            'email':            u.email,
            'telephone':        u.telephone,
            'ville':            u.ville,
            'date_inscription': u.date_inscription.isoformat() if u.date_inscription else None,
            'is_staff':         u.is_staff,
        } for u in users]
        return Response(data)


# ─── Newsletter ───────────────────────────────────────────────────────────────

from .models import NewsletterAbonne
from .serializers import NewsletterSerializer


class NewsletterInscriptionView(APIView):
    permission_classes = [permissions.AllowAny]
    throttle_scope     = 'contact'

    def post(self, request):
        email = request.data.get('email', '').strip().lower()
        if not email or '@' not in email:
            return Response({'detail': 'Email invalide.'}, status=400)

        # Vérification blacklist
        if Blacklist.objects.filter(valeur=email, type_blacklist='email').exists():
            return Response({'detail': "Cet email ne peut pas s'inscrire."}, status=403)

        abonne, created = NewsletterAbonne.objects.get_or_create(email=email)
        if not created and not abonne.actif:
            abonne.actif = True
            abonne.save()
            created = True

        if created:
            send_mail(
                subject='🌿 Bienvenue dans la famille Tropicana Pio Pio !',
                message=(
                    f"Bonjour,\n\nMerci de vous être inscrit à notre newsletter !\n\n"
                    f"Vous recevrez :\n  • Nos conseils santé et bien-être\n"
                    f"  • Nos offres exclusives\n  • Les nouveautés Tropicana Pio Pio\n\n"
                    f"Boutique : https://tropicanapiopio.com/boutique\n\n"
                    f"— L'équipe Tropicana Pio Pio 🌿"
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=True,
            )

        return Response(
            {'detail': 'Inscription confirmée ! Vérifiez votre boîte email.'},
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK
        )


class NewsletterDesabonnementView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        email = request.data.get('email', '').strip().lower()
        try:
            abonne = NewsletterAbonne.objects.get(email=email)
            abonne.actif = False
            abonne.save()
            return Response({'detail': 'Désinscription effectuée.'})
        except NewsletterAbonne.DoesNotExist:
            return Response({'detail': 'Email introuvable.'}, status=404)


class AdminNewsletterView(APIView):
    """GET : liste abonnés. POST : envoyer un email à tous les abonnés actifs."""
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        abonnes = NewsletterAbonne.objects.filter(actif=True).values('id', 'email', 'date_inscription')
        return Response({'count': abonnes.count(), 'abonnes': list(abonnes)})

    def post(self, request):
        """Envoyer une newsletter à tous les abonnés actifs."""
        sujet   = request.data.get('sujet', '').strip()
        message = request.data.get('message', '').strip()
        if not sujet or not message:
            return Response({'detail': 'Sujet et message requis.'}, status=400)

        emails = list(NewsletterAbonne.objects.filter(actif=True).values_list('email', flat=True))
        if not emails:
            return Response({'detail': 'Aucun abonné actif.'}, status=400)

        # Envoi par batch de 50
        envoyes = 0
        erreurs = 0
        batch_size = 50
        for i in range(0, len(emails), batch_size):
            batch = emails[i:i + batch_size]
            try:
                send_mail(
                    subject=f'[Tropicana Pio Pio] {sujet}',
                    message=message + '\n\n---\nPour vous désabonner : https://tropicanapiopio.com/newsletter/desabonnement',
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=batch,
                    fail_silently=False,
                )
                envoyes += len(batch)
            except Exception as e:
                erreurs += len(batch)
                logger_securite.error(f'Newsletter batch erreur : {e}')

        return Response({
            'detail':  f'Newsletter envoyée à {envoyes} abonné(s).',
            'envoyes': envoyes,
            'erreurs': erreurs,
        })


# ─── CRUD Slider, Bienfait, Partenaire, Histoire, Blog ───────────────────────

from .models import Slider, Bienfait, Partenaire, HistoireChapitre, ArticleBlog
from .serializers import (
    SliderSerializer, BienfaitSerializer, PartenaireSerializer,
    HistoireChapitreSerializer, ArticleBlogSerializer,
)


class SlidersPublicView(generics.ListAPIView):
    serializer_class   = SliderSerializer
    permission_classes = [permissions.AllowAny]
    def get_queryset(self):
        return Slider.objects.filter(actif=True).order_by('ordre')
    def get_serializer_context(self):
        ctx = super().get_serializer_context(); ctx['request'] = self.request; return ctx


class AdminSlidersView(generics.ListCreateAPIView):
    serializer_class   = SliderSerializer
    permission_classes = [permissions.IsAdminUser]
    parser_classes     = [MultiPartParser, FormParser, JSONParser]
    def get_queryset(self):
        return Slider.objects.all().order_by('ordre')
    def get_serializer_context(self):
        ctx = super().get_serializer_context(); ctx['request'] = self.request; return ctx


class AdminSliderDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class   = SliderSerializer
    permission_classes = [permissions.IsAdminUser]
    parser_classes     = [MultiPartParser, FormParser, JSONParser]
    queryset           = Slider.objects.all()
    def get_serializer_context(self):
        ctx = super().get_serializer_context(); ctx['request'] = self.request; return ctx


class BienfaitsPublicView(generics.ListAPIView):
    serializer_class   = BienfaitSerializer
    permission_classes = [permissions.AllowAny]
    def get_queryset(self):
        return Bienfait.objects.filter(actif=True).order_by('ordre')


class AdminBienfaitsView(generics.ListCreateAPIView):
    serializer_class   = BienfaitSerializer
    permission_classes = [permissions.IsAdminUser]
    def get_queryset(self):
        return Bienfait.objects.all().order_by('ordre')


class AdminBienfaitDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class   = BienfaitSerializer
    permission_classes = [permissions.IsAdminUser]
    queryset           = Bienfait.objects.all()


class PartenairesPublicView(generics.ListAPIView):
    serializer_class   = PartenaireSerializer
    permission_classes = [permissions.AllowAny]
    def get_queryset(self):
        return Partenaire.objects.filter(actif=True).order_by('ordre')
    def get_serializer_context(self):
        ctx = super().get_serializer_context(); ctx['request'] = self.request; return ctx


class AdminPartenairesView(generics.ListCreateAPIView):
    serializer_class   = PartenaireSerializer
    permission_classes = [permissions.IsAdminUser]
    parser_classes     = [MultiPartParser, FormParser, JSONParser]
    def get_queryset(self):
        return Partenaire.objects.all().order_by('ordre')
    def get_serializer_context(self):
        ctx = super().get_serializer_context(); ctx['request'] = self.request; return ctx


class AdminPartenaireDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class   = PartenaireSerializer
    permission_classes = [permissions.IsAdminUser]
    parser_classes     = [MultiPartParser, FormParser, JSONParser]
    queryset           = Partenaire.objects.all()
    def get_serializer_context(self):
        ctx = super().get_serializer_context(); ctx['request'] = self.request; return ctx


class HistoirePublicView(generics.ListAPIView):
    serializer_class   = HistoireChapitreSerializer
    permission_classes = [permissions.AllowAny]
    def get_queryset(self):
        return HistoireChapitre.objects.filter(actif=True).order_by('ordre')
    def get_serializer_context(self):
        ctx = super().get_serializer_context(); ctx['request'] = self.request; return ctx


class AdminHistoireView(generics.ListCreateAPIView):
    serializer_class   = HistoireChapitreSerializer
    permission_classes = [permissions.IsAdminUser]
    parser_classes     = [MultiPartParser, FormParser, JSONParser]
    def get_queryset(self):
        return HistoireChapitre.objects.all().order_by('ordre')
    def get_serializer_context(self):
        ctx = super().get_serializer_context(); ctx['request'] = self.request; return ctx


class AdminHistoireDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class   = HistoireChapitreSerializer
    permission_classes = [permissions.IsAdminUser]
    parser_classes     = [MultiPartParser, FormParser, JSONParser]
    queryset           = HistoireChapitre.objects.all()
    def get_serializer_context(self):
        ctx = super().get_serializer_context(); ctx['request'] = self.request; return ctx


class BlogPublicView(generics.ListAPIView):
    serializer_class   = ArticleBlogSerializer
    permission_classes = [permissions.AllowAny]
    def get_queryset(self):
        return ArticleBlog.objects.filter(publie=True).order_by('-date_publication')
    def get_serializer_context(self):
        ctx = super().get_serializer_context(); ctx['request'] = self.request; return ctx


class BlogDetailPublicView(generics.RetrieveAPIView):
    serializer_class   = ArticleBlogSerializer
    permission_classes = [permissions.AllowAny]
    lookup_field       = 'slug'
    def get_queryset(self):
        return ArticleBlog.objects.filter(publie=True)
    def get_serializer_context(self):
        ctx = super().get_serializer_context(); ctx['request'] = self.request; return ctx


class AdminBlogView(generics.ListCreateAPIView):
    serializer_class   = ArticleBlogSerializer
    permission_classes = [permissions.IsAdminUser]
    parser_classes     = [MultiPartParser, FormParser, JSONParser]
    def get_queryset(self):
        return ArticleBlog.objects.all().order_by('-date_publication')
    def get_serializer_context(self):
        ctx = super().get_serializer_context(); ctx['request'] = self.request; return ctx


class AdminBlogDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class   = ArticleBlogSerializer
    permission_classes = [permissions.IsAdminUser]
    parser_classes     = [MultiPartParser, FormParser, JSONParser]
    queryset           = ArticleBlog.objects.all()
    def get_serializer_context(self):
        ctx = super().get_serializer_context(); ctx['request'] = self.request; return ctx


# ─── Mission, Fondateur, ConfigAccueil, ConfigSite, FAQ ──────────────────────

from .models import Mission, FondateurConfig
from .serializers import MissionSerializer, FondateurConfigSerializer


class MissionsPublicView(generics.ListAPIView):
    serializer_class   = MissionSerializer
    permission_classes = [permissions.AllowAny]
    def get_queryset(self):
        return Mission.objects.filter(actif=True).order_by('ordre')


class AdminMissionsView(generics.ListCreateAPIView):
    serializer_class   = MissionSerializer
    permission_classes = [permissions.IsAdminUser]
    def get_queryset(self):
        return Mission.objects.all().order_by('ordre')


class AdminMissionDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class   = MissionSerializer
    permission_classes = [permissions.IsAdminUser]
    queryset           = Mission.objects.all()


class FondateurPublicView(generics.RetrieveAPIView):
    serializer_class   = FondateurConfigSerializer
    permission_classes = [permissions.AllowAny]
    def get_object(self):
        obj, _ = FondateurConfig.objects.get_or_create(pk=1); return obj
    def get_serializer_context(self):
        ctx = super().get_serializer_context(); ctx['request'] = self.request; return ctx


class AdminFondateurView(generics.RetrieveUpdateAPIView):
    serializer_class   = FondateurConfigSerializer
    permission_classes = [permissions.IsAdminUser]
    parser_classes     = [MultiPartParser, FormParser, JSONParser]
    def get_object(self):
        obj, _ = FondateurConfig.objects.get_or_create(pk=1); return obj
    def get_serializer_context(self):
        ctx = super().get_serializer_context(); ctx['request'] = self.request; return ctx


from .models import ConfigAccueil
from .serializers import ConfigAccueilSerializer


class ConfigAccueilPublicView(generics.RetrieveAPIView):
    serializer_class   = ConfigAccueilSerializer
    permission_classes = [permissions.AllowAny]
    def get_object(self):
        obj, _ = ConfigAccueil.objects.get_or_create(pk=1); return obj
    def get_serializer_context(self):
        ctx = super().get_serializer_context(); ctx['request'] = self.request; return ctx


class AdminConfigAccueilView(generics.RetrieveUpdateAPIView):
    serializer_class   = ConfigAccueilSerializer
    permission_classes = [permissions.IsAdminUser]
    parser_classes     = [MultiPartParser, FormParser, JSONParser]
    def get_object(self):
        obj, _ = ConfigAccueil.objects.get_or_create(pk=1); return obj
    def get_serializer_context(self):
        ctx = super().get_serializer_context(); ctx['request'] = self.request; return ctx


from .models import ConfigSite
from .serializers import ConfigSiteSerializer


class ConfigSitePublicView(generics.RetrieveAPIView):
    serializer_class   = ConfigSiteSerializer
    permission_classes = [permissions.AllowAny]
    def get_object(self):
        obj, _ = ConfigSite.objects.get_or_create(pk=1); return obj


class AdminConfigSiteView(generics.RetrieveUpdateAPIView):
    serializer_class   = ConfigSiteSerializer
    permission_classes = [permissions.IsAdminUser]
    def get_object(self):
        obj, _ = ConfigSite.objects.get_or_create(pk=1); return obj


from .models import FAQ
from .serializers import FAQSerializer


class FAQPublicView(generics.ListAPIView):
    serializer_class   = FAQSerializer
    permission_classes = [permissions.AllowAny]
    def get_queryset(self):
        return FAQ.objects.filter(actif=True).order_by('ordre')


class AdminFAQView(generics.ListCreateAPIView):
    serializer_class   = FAQSerializer
    permission_classes = [permissions.IsAdminUser]
    def get_queryset(self):
        return FAQ.objects.all().order_by('ordre')


class AdminFAQDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class   = FAQSerializer
    permission_classes = [permissions.IsAdminUser]
    queryset           = FAQ.objects.all()


# ─── Logs de connexion ────────────────────────────────────────────────────────

class AdminLogsConnexionView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        page      = int(request.query_params.get('page', 1))
        per_page  = int(request.query_params.get('per_page', 50))
        resultat  = request.query_params.get('resultat')
        email_q   = request.query_params.get('email')

        qs = LogConnexion.objects.all()
        if resultat:
            qs = qs.filter(resultat=resultat)
        if email_q:
            qs = qs.filter(email__icontains=email_q)

        total = qs.count()
        start = (page - 1) * per_page
        logs  = qs[start:start + per_page]

        data = [{
            'id':         l.id,
            'email':      l.email,
            'ip':         l.ip,
            'resultat':   l.resultat,
            'date':       l.date.isoformat(),
            'user_agent': l.user_agent[:100],
        } for l in logs]

        hier = timezone.now() - datetime.timedelta(hours=24)
        stats = {
            'total_24h':  LogConnexion.objects.filter(date__gte=hier).count(),
            'echecs_24h': LogConnexion.objects.filter(date__gte=hier, resultat='echec').count(),
            'ips_suspectes': list(
                LogConnexion.objects.filter(date__gte=hier, resultat='echec')
                .values('ip').annotate(nb=models.Count('ip')).filter(nb__gte=5)
                .values_list('ip', flat=True)
            ),
        }

        return Response({
            'count':   total,
            'page':    page,
            'pages':   (total + per_page - 1) // per_page,
            'results': data,
            'stats':   stats,
        })


# ═══════════════════════════════════════════════════════════════════════════════
# NOUVELLES FONCTIONNALITÉS
# ═══════════════════════════════════════════════════════════════════════════════

from .models import (
    HistoriqueCommande, ZoneLivraison, Blacklist, ReponseAvis, AlerteStock
)


# ─── Zones de livraison ───────────────────────────────────────────────────────

class ZonesLivraisonPublicView(APIView):
    """Liste les zones de livraison disponibles avec leurs prix."""
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        from .models import ZoneLivraison
        zones = ZoneLivraison.objects.filter(disponible=True).order_by('ordre')
        data = [{'ville': z.ville, 'prix': int(z.prix), 'delai': z.delai} for z in zones]
        return Response(data)


class AdminZonesLivraisonView(APIView):
    """CRUD zones de livraison — admin seulement."""
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        zones = ZoneLivraison.objects.all().order_by('ordre')
        data = [{'id': z.id, 'ville': z.ville, 'prix': int(z.prix),
                 'delai': z.delai, 'disponible': z.disponible, 'ordre': z.ordre}
                for z in zones]
        return Response(data)

    def post(self, request):
        ville = request.data.get('ville', '').strip()
        if not ville:
            return Response({'detail': 'Ville requise.'}, status=400)
        if ZoneLivraison.objects.filter(ville__iexact=ville).exists():
            return Response({'detail': 'Cette ville existe déjà.'}, status=400)
        z = ZoneLivraison.objects.create(
            ville=ville,
            prix=request.data.get('prix', 0),
            delai=request.data.get('delai', '24-48h'),
            disponible=request.data.get('disponible', True),
            ordre=request.data.get('ordre', 0),
        )
        return Response({'detail': 'Zone créée.', 'id': z.id}, status=201)


class AdminZoneLivraisonDetailView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def patch(self, request, pk):
        try:
            z = ZoneLivraison.objects.get(pk=pk)
        except ZoneLivraison.DoesNotExist:
            return Response({'detail': 'Introuvable.'}, status=404)
        for champ in ['ville', 'prix', 'delai', 'disponible', 'ordre']:
            if champ in request.data:
                setattr(z, champ, request.data[champ])
        z.save()
        return Response({'detail': 'Mis à jour.'})

    def delete(self, request, pk):
        ZoneLivraison.objects.filter(pk=pk).delete()
        return Response({'detail': 'Supprimé.'})


# ─── Historique commande ──────────────────────────────────────────────────────

class AdminHistoriqueCommandeView(APIView):
    """Historique des changements de statut d'une commande."""
    permission_classes = [permissions.IsAdminUser]

    def get(self, request, pk):
        histo = HistoriqueCommande.objects.filter(commande_id=pk).order_by('-date')
        data = [{
            'id':             h.id,
            'ancien_statut':  h.ancien_statut,
            'nouveau_statut': h.nouveau_statut,
            'note':           h.note,
            'modifie_par':    h.modifie_par.email if h.modifie_par else 'Système',
            'date':           h.date.isoformat(),
        } for h in histo]
        return Response(data)


# ─── Blacklist ────────────────────────────────────────────────────────────────

class AdminBlacklistView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        bl = Blacklist.objects.all().order_by('-date_ajout')
        data = [{'id': b.id, 'type': b.type_blacklist, 'valeur': b.valeur,
                 'raison': b.raison, 'date': b.date_ajout.isoformat()} for b in bl]
        return Response({'count': len(data), 'results': data})

    def post(self, request):
        valeur = request.data.get('valeur', '').strip().lower()
        type_bl = request.data.get('type_blacklist', 'email')
        bl, created = Blacklist.objects.get_or_create(
            valeur=valeur,
            defaults={
                'type_blacklist': type_bl,
                'raison':         request.data.get('raison', ''),
                'ajoute_par':     request.user,
            }
        )
        if not created:
            return Response({'detail': 'Déjà en blacklist.'}, status=400)
        return Response({'detail': 'Ajouté à la blacklist.', 'id': bl.id}, status=201)


class AdminBlacklistDetailView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def delete(self, request, pk):
        Blacklist.objects.filter(pk=pk).delete()
        return Response({'detail': 'Retiré de la blacklist.'})


# ─── Réponse aux avis ─────────────────────────────────────────────────────────

class AdminReponseAvisView(APIView):
    """Ajouter ou modifier la réponse de l'admin à un témoignage."""
    permission_classes = [permissions.IsAdminUser]

    def post(self, request, pk):
        try:
            temoignage = Temoignage.objects.get(pk=pk)
        except Temoignage.DoesNotExist:
            return Response({'detail': 'Avis introuvable.'}, status=404)
        texte = request.data.get('texte', '').strip()
        if not texte:
            return Response({'detail': 'Texte requis.'}, status=400)
        reponse, created = ReponseAvis.objects.update_or_create(
            temoignage=temoignage,
            defaults={'texte': texte, 'modifie_par': request.user}
        )
        return Response({'detail': 'Réponse enregistrée.', 'created': created})

    def delete(self, request, pk):
        ReponseAvis.objects.filter(temoignage_id=pk).delete()
        return Response({'detail': 'Réponse supprimée.'})


# ─── Alertes stock ────────────────────────────────────────────────────────────

class AdminAlertesStockView(APIView):
    """Gérer les seuils d'alerte de stock par produit."""
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        alertes = AlerteStock.objects.select_related('produit').all()
        data = [{
            'id':            a.id,
            'produit_id':    a.produit_id,
            'produit_nom':   a.produit.nom,
            'stock_actuel':  a.produit.stock,
            'seuil':         a.seuil,
            'email_alerte':  a.email_alerte,
            'derniere_alerte': a.derniere_alerte.isoformat() if a.derniere_alerte else None,
            'en_alerte':     a.produit.stock > 0 and a.produit.stock <= a.seuil,
        } for a in alertes]
        return Response(data)

    def post(self, request):
        produit_id = request.data.get('produit_id')
        seuil      = request.data.get('seuil', 5)
        try:
            produit = Produit.objects.get(pk=produit_id)
        except Produit.DoesNotExist:
            return Response({'detail': 'Produit introuvable.'}, status=404)
        alerte, _ = AlerteStock.objects.update_or_create(
            produit=produit,
            defaults={
                'seuil':        seuil,
                'email_alerte': request.data.get('email_alerte', ''),
            }
        )
        return Response({'detail': 'Alerte configurée.', 'id': alerte.id})


def verifier_alertes_stock():
    """Appelé après chaque commande — envoie un email si stock bas."""
    from django.utils import timezone
    import datetime
    alertes = AlerteStock.objects.select_related('produit').all()
    for alerte in alertes:
        p = alerte.produit
        if p.stock > 0 and p.stock <= alerte.seuil:
            # Éviter de spammer : max 1 alerte par 24h par produit
            if alerte.derniere_alerte and \
               alerte.derniere_alerte > timezone.now() - datetime.timedelta(hours=24):
                continue
            destinataire = alerte.email_alerte or settings.DEFAULT_FROM_EMAIL
            send_mail(
                subject=f'⚠️ Stock bas — {p.nom} ({p.stock} restants)',
                message=(
                    f"Attention, le stock du produit \"{p.nom}\" est bas.\n\n"
                    f"Stock actuel : {p.stock}\n"
                    f"Seuil d'alerte : {alerte.seuil}\n\n"
                    f"Pensez à réapprovisionner rapidement.\n\n"
                    f"— Système Tropicana Pio Pio"
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[destinataire],
                fail_silently=True,
            )
            alerte.derniere_alerte = timezone.now()
            alerte.save(update_fields=['derniere_alerte'])


# ─── Remboursement ────────────────────────────────────────────────────────────

class AdminRemboursementView(APIView):
    """Marquer une commande comme remboursée et notifier le client."""
    permission_classes = [permissions.IsAdminUser]

    def post(self, request, pk):
        try:
            commande = Commande.objects.get(pk=pk)
        except Commande.DoesNotExist:
            return Response({'detail': 'Commande introuvable.'}, status=404)

        note = request.data.get('note', '').strip()
        ancien_statut = commande.statut
        commande.statut = 'annulee'
        commande.notes  = (commande.notes + f'\n\n[REMBOURSEMENT] {note}').strip()
        commande.save(update_fields=['statut', 'notes'])

        # Historique
        HistoriqueCommande.objects.create(
            commande=commande,
            ancien_statut=ancien_statut,
            nouveau_statut='annulee',
            note=f'Remboursement : {note}',
            modifie_par=request.user,
        )

        # Email client
        if commande.email_client:
            send_mail(
                subject=f'[Tropicana Pio Pio] 💰 Remboursement — Commande #{commande.pk}',
                message=(
                    f"Bonjour {commande.nom_client},\n\n"
                    f"Votre commande #{commande.pk} a été remboursée.\n"
                    f"Montant : {commande.total} FCFA\n"
                    f"{('Motif : ' + note) if note else ''}\n\n"
                    f"Pour toute question :\n"
                    f"📞 +229 01 95 96 77 62\n"
                    f"💬 WhatsApp : wa.me/2290195967762\n\n"
                    f"— L'équipe Tropicana Pio Pio"
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[commande.email_client],
                fail_silently=True,
            )

        return Response({'detail': f'Commande #{pk} marquée comme remboursée. Client notifié.'})


# ─── Recherche globale admin ──────────────────────────────────────────────────

class AdminRechercheGlobaleView(APIView):
    """Recherche dans commandes + clients + messages simultanément."""
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        q = request.query_params.get('q', '').strip()
        if len(q) < 2:
            return Response({'detail': 'Requête trop courte (min 2 caractères).'}, status=400)

        commandes = Commande.objects.filter(
            models.Q(nom_client__icontains=q) |
            models.Q(email_client__icontains=q) |
            models.Q(telephone_client__icontains=q) |
            models.Q(ville_livraison__icontains=q)
        ).order_by('-date_commande')[:10]

        clients = Utilisateur.objects.filter(
            models.Q(prenom__icontains=q) |
            models.Q(nom__icontains=q) |
            models.Q(email__icontains=q) |
            models.Q(telephone__icontains=q)
        )[:10]

        messages = MessageContact.objects.filter(
            models.Q(nom__icontains=q) |
            models.Q(email__icontains=q) |
            models.Q(sujet__icontains=q) |
            models.Q(message__icontains=q)
        ).order_by('-date_envoi')[:10]

        return Response({
            'query': q,
            'commandes': [{
                'id': c.pk, 'nom': c.nom_client, 'total': int(c.total),
                'statut': c.statut, 'date': c.date_commande.strftime('%d/%m/%Y'),
            } for c in commandes],
            'clients': [{
                'id': u.pk, 'nom': f'{u.prenom} {u.nom}',
                'email': u.email, 'telephone': u.telephone,
            } for u in clients],
            'messages': [{
                'id': m.pk, 'nom': m.nom, 'email': m.email,
                'sujet': m.sujet, 'lu': m.lu,
            } for m in messages],
        })


# ─── Rapport PDF mensuel ──────────────────────────────────────────────────────

class AdminRapportPDFView(APIView):
    """Génère un rapport PDF mensuel des ventes."""
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        from django.http import HttpResponse
        from django.db.models import Sum, Count
        import io

        mois  = int(request.query_params.get('mois', timezone.now().month))
        annee = int(request.query_params.get('annee', timezone.now().year))

        commandes = Commande.objects.filter(
            date_commande__month=mois,
            date_commande__year=annee,
        ).prefetch_related('lignes')

        ca_total = commandes.filter(statut='livree').aggregate(s=Sum('total'))['s'] or 0
        nb_total = commandes.count()
        nb_livrees = commandes.filter(statut='livree').count()

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm

            buffer = io.BytesIO()
            doc    = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
            styles = getSampleStyleSheet()
            vert   = colors.HexColor('#2D6A4F')
            mois_fr = ['','Janvier','Février','Mars','Avril','Mai','Juin',
                       'Juillet','Août','Septembre','Octobre','Novembre','Décembre']
            elements = []

            # Titre
            titre_style = ParagraphStyle('titre', parent=styles['Title'], textColor=vert, fontSize=18)
            elements.append(Paragraph(f'Rapport mensuel — {mois_fr[mois]} {annee}', titre_style))
            elements.append(Paragraph('Tropicana Pio Pio', styles['Normal']))
            elements.append(Spacer(1, 0.5*cm))

            # Résumé
            resume_data = [
                ['Indicateur', 'Valeur'],
                ['Total commandes', str(nb_total)],
                ['Commandes livrées', str(nb_livrees)],
                ['Chiffre d\'affaires (livrées)', f'{int(ca_total):,} FCFA'.replace(',', ' ')],
            ]
            t = Table(resume_data, colWidths=[10*cm, 7*cm])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), vert),
                ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
                ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
                ('GRID',       (0,0), (-1,-1), 0.5, colors.lightgrey),
                ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F8FAF8')),
                ('FONTSIZE',   (0,0), (-1,-1), 11),
                ('PADDING',    (0,0), (-1,-1), 8),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 0.7*cm))

            # Tableau commandes
            elements.append(Paragraph('Détail des commandes', ParagraphStyle('h2', parent=styles['Heading2'], textColor=vert)))
            elements.append(Spacer(1, 0.3*cm))
            rows = [['#', 'Date', 'Client', 'Ville', 'Total', 'Statut']]
            for c in commandes.order_by('-date_commande'):
                rows.append([
                    str(c.pk),
                    c.date_commande.strftime('%d/%m/%Y'),
                    c.nom_client[:25],
                    c.ville_livraison[:20],
                    f'{int(c.total):,} FCFA'.replace(',', ' '),
                    c.get_statut_display(),
                ])
            t2 = Table(rows, colWidths=[1.2*cm, 2.5*cm, 5*cm, 3.5*cm, 3.5*cm, 3*cm])
            t2.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), vert),
                ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
                ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
                ('GRID',       (0,0), (-1,-1), 0.3, colors.lightgrey),
                ('FONTSIZE',   (0,0), (-1,-1), 8),
                ('PADDING',    (0,0), (-1,-1), 5),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F0F4F1')]),
            ]))
            elements.append(t2)

            doc.build(elements)
            buffer.seek(0)
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="rapport_{mois_fr[mois]}_{annee}.pdf"'
            return response

        except ImportError:
            # reportlab non installé → rapport texte simple
            lines = [
                f'RAPPORT MENSUEL — {mois}/{annee}',
                f'Tropicana Pio Pio',
                '=' * 40,
                f'Total commandes : {nb_total}',
                f'Commandes livrées : {nb_livrees}',
                f'CA total : {int(ca_total):,} FCFA'.replace(',', ' '),
                '',
                'DÉTAIL :',
            ]
            for c in commandes:
                lines.append(f'#{c.pk} | {c.date_commande.strftime("%d/%m/%Y")} | {c.nom_client} | {int(c.total)} FCFA | {c.get_statut_display()}')
            response = HttpResponse('\n'.join(lines), content_type='text/plain; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="rapport_{mois}_{annee}.txt"'
            return response


# ─── Impression bon de commande ───────────────────────────────────────────────

class AdminBonCommandeView(APIView):
    """Génère un bon de commande imprimable (HTML)."""
    permission_classes = [permissions.IsAdminUser]

    def get(self, request, pk):
        from django.http import HttpResponse
        try:
            commande = Commande.objects.prefetch_related('lignes__produit').get(pk=pk)
        except Commande.DoesNotExist:
            return Response({'detail': 'Commande introuvable.'}, status=404)

        lignes_html = ''.join([
            f'<tr><td>{l.produit.nom if l.produit else "Produit"}</td><td style="text-align:center">{l.quantite}</td>'
            f'<td style="text-align:right">{int(l.prix_unitaire if hasattr(l, "prix_unitaire") else "" if hasattr(l, "prix_unitaire") else ""):,} FCFA</td>'
            f'<td style="text-align:right">{int(l.sous_total):,} FCFA</td></tr>'
            for l in commande.lignes.all()
        ])

        html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<title>Bon de commande #{commande.pk}</title>
<style>
  body {{ font-family: Arial, sans-serif; max-width: 800px; margin: 0 auto; padding: 20px; color: #333; }}
  .header {{ display: flex; justify-content: space-between; align-items: center; border-bottom: 3px solid #2D6A4F; padding-bottom: 16px; margin-bottom: 24px; }}
  .logo {{ font-size: 1.5rem; font-weight: 900; color: #2D6A4F; }}
  .badge {{ background: #2D6A4F; color: white; padding: 6px 16px; border-radius: 20px; font-size: .85rem; }}
  .info-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 24px; }}
  .info-box {{ background: #F0F4F1; padding: 14px; border-radius: 8px; }}
  .info-box h3 {{ margin: 0 0 8px; color: #2D6A4F; font-size: .85rem; text-transform: uppercase; }}
  .info-box p {{ margin: 4px 0; font-size: .9rem; }}
  table {{ width: 100%; border-collapse: collapse; margin-bottom: 16px; }}
  th {{ background: #2D6A4F; color: white; padding: 10px; text-align: left; font-size: .85rem; }}
  td {{ padding: 10px; border-bottom: 1px solid #eee; font-size: .9rem; }}
  tr:nth-child(even) td {{ background: #F8FAF8; }}
  .total {{ text-align: right; font-size: 1.2rem; font-weight: 700; color: #2D6A4F; padding: 12px 0; }}
  .footer {{ margin-top: 32px; text-align: center; color: #999; font-size: .8rem; border-top: 1px solid #eee; padding-top: 12px; }}
  @media print {{ .no-print {{ display: none; }} }}
</style>
</head>
<body>
<div class="no-print" style="text-align:right;margin-bottom:12px;">
  <button onclick="window.print()" style="background:#2D6A4F;color:white;border:none;padding:10px 24px;border-radius:8px;font-size:1rem;cursor:pointer;">🖨️ Imprimer</button>
</div>
<div class="header">
  <div class="logo">🌿 Tropicana Pio Pio</div>
  <div>
    <div class="badge">Commande #{commande.pk}</div>
    <div style="text-align:right;margin-top:6px;color:#666;font-size:.85rem;">{commande.date_commande.strftime('%d/%m/%Y à %H:%M')}</div>
  </div>
</div>
<div class="info-grid">
  <div class="info-box">
    <h3>👤 Client</h3>
    <p><strong>{commande.nom_client}</strong></p>
    <p>📞 {commande.telephone_client}</p>
    <p>✉️ {commande.email_client}</p>
  </div>
  <div class="info-box">
    <h3>📦 Livraison</h3>
    <p><strong>{commande.ville_livraison}</strong></p>
    <p>{commande.adresse_livraison or 'Adresse non précisée'}</p>
    <p>💳 {commande.get_mode_paiement_display()}</p>
    <p>Statut : <strong>{commande.get_statut_display()}</strong></p>
  </div>
</div>
<table>
  <thead><tr><th>Produit</th><th style="text-align:center">Qté</th><th style="text-align:right">Prix unit.</th><th style="text-align:right">Sous-total</th></tr></thead>
  <tbody>{lignes_html}</tbody>
</table>
<div class="total">Total : {int(commande.total):,} FCFA</div>
{f'<p style="color:#2D6A4F;font-size:.85rem;">Code promo appliqué : {commande.code_promo}</p>' if commande.code_promo else ''}
{f'<p style="color:#666;font-size:.85rem;font-style:italic;">Notes : {commande.notes}</p>' if commande.notes else ''}
<div class="footer">
  Tropicana Pio Pio · Porto-Novo, Bénin · +229 01 95 96 77 62 · tropicanapiopio.com
</div>
</body>
</html>"""
        return HttpResponse(html, content_type='text/html; charset=utf-8')


# ─── Polling nouvelles commandes (toutes les 30s) ─────────────────────────────

class AdminNouvellesCommandesView(APIView):
    """Retourne le nombre de commandes en attente depuis X secondes — pour le polling."""
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        depuis = request.query_params.get('depuis')
        qs = Commande.objects.filter(statut='en_attente')
        if depuis:
            try:
                from datetime import datetime
                dt = datetime.fromisoformat(depuis)
                qs = qs.filter(date_commande__gt=dt)
            except ValueError:
                pass
        nouvelles = qs.count()
        derniere  = Commande.objects.order_by('-date_commande').first()
        return Response({
            'nouvelles_commandes': nouvelles,
            'derniere_date': derniere.date_commande.isoformat() if derniere else None,
            'timestamp': timezone.now().isoformat(),
        })


# ─── Suivi commande public ────────────────────────────────────────────────────

class SuiviCommandeView(APIView):
    """Permet à un client de suivre sa commande sans être connecté (id + email)."""
    permission_classes = [permissions.AllowAny]
    throttle_scope     = 'contact'

    def get(self, request):
        commande_id = request.query_params.get('id', '').strip()
        email       = request.query_params.get('email', '').strip().lower()

        if not commande_id or not email:
            return Response({'detail': 'Numéro de commande et email requis.'}, status=400)

        try:
            commande = Commande.objects.prefetch_related('lignes__produit').get(
                pk=commande_id,
                email_client__iexact=email,
            )
        except Commande.DoesNotExist:
            return Response({'detail': 'Commande introuvable. Vérifiez votre numéro et votre email.'}, status=404)

        lignes = [{
            'produit_nom':  l.produit.nom if l.produit else 'Produit supprimé',
            'quantite':     l.quantite,
            'prix_unitaire': int(l.prix_unitaire if hasattr(l, "prix_unitaire") else "" if hasattr(l, "prix_unitaire") else ""),
            'sous_total':   int(l.sous_total),
        } for l in commande.lignes.all()]

        return Response({
            'id':               commande.pk,
            'statut':           commande.statut,
            'nom_client':       commande.nom_client,
            'telephone_client': commande.telephone_client,
            'ville_livraison':  commande.ville_livraison,
            'adresse_livraison': commande.adresse_livraison,
            'total':            int(commande.total),
            'mode_paiement':    commande.get_mode_paiement_display(),
            'payee':            commande.payee,
            'date_commande':    commande.date_commande.isoformat(),
            'lignes':           lignes,
        })
